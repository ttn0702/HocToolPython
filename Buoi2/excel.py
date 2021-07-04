from openpyxl import load_workbook
def read_cell(file_path,sheetname,cell_name):
    wb=load_workbook(filename =file_path)
    sheet_ranges = wb[sheetname]
    return sheet_ranges[cell_name].value
def update_cell(file_path,sheetname,cell_name,new_value):
    wb=load_workbook(filename=file_path)
    wb[sheetname][cell_name].value = new_value
    wb.close()
    wb.save(file_path)
if __name__ == "__main__":
    L1 = ["Nguyen Ba Nghia",
        "Nguyen Van Nam",
        "Nguyen Nhu Thuy"
    ]
    L2 = [29,23,42]
    update_cell('test.xlsx','Sheet1','A1','Họ và tên')
    update_cell('test.xlsx','Sheet1','B1','Tuổi')
    for i in range(0,len(L1)):
        name = L1[i]
        age = L2[i]
        print( name,age)